import requests
from bs4 import BeautifulSoup
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

urls = {
    'za-volontere': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-volontere/271',
    'za-javnost': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-javnost/272',
    'za-djelatnike': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-djelatnike/270'
}

def get_education_data(link):
    response = requests.get(link)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')

    education_data = {}

    try:
        title = soup.find('h1').text.strip()
    except:
        title = ''
    education_data['Title'] = title

    try:
        description_div = soup.find('div', class_='page-text')
        description = ''
        inside_description = True

        for element in description_div.children:
            if element.name == 'strong':
                strong_text = element.text.strip()

                if ':' in strong_text:
                    inside_description = False

            if inside_description:
                if isinstance(element, str):
                    description += element.strip() + ' '
                elif element.name == 'strong':
                    description += element.text.strip() + ' '
            else:
                break

        description = description.strip()
    except Exception as e:
        print(f"Error parsing description: {e}")
        description = ''
    education_data['Description'] = description

    try:
        for strong in description_div.find_all('strong'):
            label = strong.text.strip().replace(':', '').lower()
            if ':' in strong.text:
                value = ''
                next_sibling = strong.next_sibling
                while next_sibling:
                    if isinstance(next_sibling, str):
                        value += next_sibling.strip()

                    elif next_sibling.name == 'ul':
                        for li in next_sibling.find_all('li'):
                            value += li.text.strip() + '; '

                    elif next_sibling.name == 'br':
                        value += ' '

                    if next_sibling.name == 'strong':
                        break
                    next_sibling = next_sibling.next_sibling

                education_data[label.capitalize()] = value.strip('; \n')
    except Exception as e:
        print(f"Error parsing details: {e}")

    return education_data

def get_educations(url):
    data = []
    page = 1

    while True:
        response = requests.get(f"{url}?page={page}")
        response.encoding = 'utf-8' 
        soup = BeautifulSoup(response.text, 'html.parser')

        educations = soup.find_all('div', class_='block-standard publikacije')
        
        if not educations:
            break

        for education in educations:
            title_tag = education.find('div', class_='bs-title')
            if title_tag:
                title = title_tag.text.strip()
                link_tag = title_tag.find('a')
                if link_tag:
                    link = 'https://www.hck.hr' + link_tag['href']
                    details = get_education_data(link)
                    data.append(details)

        pagination = soup.find('ul', class_='cpagination')
        if pagination and pagination.find('li', class_='active'):
            active_page = pagination.find('li', class_='active').text
            if int(active_page) == page:
                page += 1  
            else:
                break
        else:
            break

    return data

def write_to_excel(workbook, category, data):
    df = pd.DataFrame(data)
    ws = workbook.create_sheet(title=category)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

wb = Workbook()

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

for category, url in urls.items():
    data = get_educations(url)
    write_to_excel(wb, category, data)

wb.save("edukacije.xlsx")
print("Data saved to edukacije.xlsx")
