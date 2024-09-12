import requests
from bs4 import BeautifulSoup
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

urls = {
    'za-volontere': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-volontere/271',
    'za-javnost': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-javnost/272',
    'za-djelatnike': 'https://www.hck.hr/edukacije-publikacije/edukacije-hrvatskog-crvenog-kriza/za-djelatnike/270'
}

def get_education_data(link):
    response = requests.get(link)
    response.encoding = 'utf-8'  
    soup = BeautifulSoup(response.text, 'html.parser')

    try:
        title = soup.find('h1').text.strip()
    except:
        title = ''

    try:
        description_div = soup.find('div', class_='page-text')
        description = ''
        for elem in description_div.stripped_strings: 
            if elem.startswith('Preduvjet: ') or elem.startswith('Predavači: ') or elem.startswith('Trajanje: '):
                break
            description += elem + ' '
            
        description = description.strip()
    except:
        description = ''

    details = soup.find('div', class_='page-text')
    
    preduvjet = ''
    predavaci = ''
    trajanje = ''

    for strong in details.find_all('strong'):
        label = strong.text.strip().lower()
        if 'preduvjet' in label:
            preduvjet = strong.next_sibling.strip() if strong.next_sibling else ''
        elif 'predavači' in label:
            predavaci = strong.next_sibling.strip() if strong.next_sibling else ''
        elif 'trajanje' in label:
            trajanje = strong.next_sibling.strip() if strong.next_sibling else ''

    return {
        'Title': title,
        'Description': description,
        'Preduvjet': preduvjet,
        'Predavači': predavaci,
        'Trajanje': trajanje
    }

def get_educations(url):
    data = []
    page = 1  

    while True:
        response = requests.get(f"{url}?page={page}")
        response.encoding = 'utf-8' 
        soup = BeautifulSoup(response.text, 'html.parser')

        educations = soup.find_all('div', class_='block-standard publikacije')
        
        if not educations:
            break

        for education in educations:
            title_tag = education.find('div', class_='bs-title')
            if title_tag:
                title = title_tag.text.strip()
                link_tag = title_tag.find('a')
                if link_tag:
                    link = 'https://www.hck.hr' + link_tag['href']
                    details = get_education_data(link)
                    data.append(details)

        pagination = soup.find('ul', class_='cpagination')
        if pagination and pagination.find('li', class_='active'):
            active_page = pagination.find('li', class_='active').text
            if int(active_page) == page:
                page += 1  
            else:
                break
        else:
            break

    return data

def write_to_excel(workbook, category, data):
    ws = workbook.create_sheet(title=category)

    df = pd.DataFrame(data)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True)  
        cell.alignment = Alignment(horizontal="center", vertical="center")
    


wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

for category, url in urls.items():
    data = get_educations(url)
    write_to_excel(wb, category, data)

wb.save("edukacije.xlsx")
print("Data saved to edukacije.xlsx")

